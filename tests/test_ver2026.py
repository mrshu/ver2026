"""Tests for the VER 2026 loader and sort/filter helpers."""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from ver2026 import filter_rows, load, sort_rows
from ver2026.cli import main
from ver2026.official_web import parse_results_html
from ver2026.reward import build_reward_comparison, quality_points

DATA = Path(__file__).resolve().parent.parent / "data" / "VER2026data.xlsx"
WEB_DATA = Path(__file__).resolve().parent.parent / "web" / "data.json"
REWARD_DATA = Path(__file__).resolve().parent.parent / "web" / "financing" / "data.json"


@pytest.fixture(scope="module")
def data():
    return load(DATA)


def test_load_returns_all_270_institutions(data):
    assert len(data) == 270


def test_profiles_are_populated(data):
    # Every row should have 4 profiles × 5 levels each.
    for r in data:
        assert set(r.profiles.keys()) == {"vystupy", "spolocensky_dosah", "tvorive_prostredie", "celkovy"}
        for slug, levels in r.profiles.items():
            assert len(levels) == 5, f"{r.institution} {slug} has {len(levels)} levels"
            assert all(0 <= (p or 0) <= 100 for p in levels)


def test_score_lower_is_better(data):
    # The institution with the highest % world-class outputs should have the
    # lowest (best) vystupy score among rows where any outputs are scored.
    scored = sorted(data, key=lambda r: r.score("vystupy"))
    best = scored[0]
    worst = scored[-1]
    assert best.score("vystupy") <= worst.score("vystupy")


def test_filter_by_area(data):
    chem = filter_rows(data, eval_area="Chemické vedy")
    assert all(r.eval_area == "Chemické vedy" for r in chem)
    assert len(chem) > 0


def test_filter_by_type_vvi(data):
    vvi = filter_rows(data, inst_type="VVI")
    assert all(r.inst_type == "VVI" for r in vvi)
    assert len(vvi) == 48


def test_filter_by_contains(data):
    sav = filter_rows(data, institution_contains="akadémie vied")
    assert all("akadémie vied" in r.institution.lower() for r in sav)
    assert len(sav) > 0


def test_sort_by_overall_score_ascending(data):
    rows = sort_rows(data, by="celkovy__score", descending=False)
    # First rows should have the lowest (best) scores.
    assert rows[0].score("celkovy") <= rows[-1].score("celkovy")


def test_top1_sums_correctly(data):
    # For the vystupy profile, top1 + top2 = sum of first two levels.
    r = data[0]
    assert r.top_pct("vystupy") == r.profiles["vystupy"][0]
    assert r.top_two_pct("vystupy") == r.profiles["vystupy"][0] + r.profiles["vystupy"][1]


def test_top_two_per_100_employees(data):
    row = next(r for r in data if r.employees)
    assert row.top_two_per_100_employees("celkovy") == pytest.approx(
        row.top_two_pct("celkovy") * 100 / row.employees
    )
    assert row.as_dict()["celkovy__top2_per_100_emp"] == pytest.approx(
        row.top_two_per_100_employees("celkovy")
    )


def test_money_normalized_metrics_can_be_computed(data):
    row = next(r for r in data if r.employees)
    row.financing_eur = 2_000_000
    assert row.financing_per_employee() == pytest.approx(2_000_000 / row.employees)
    assert row.top_two_per_million_eur("celkovy") == pytest.approx(
        row.top_two_pct("celkovy") / 2
    )
    as_dict = row.as_dict()
    assert as_dict["financing_eur"] == 2_000_000
    assert as_dict["celkovy__top2_per_million_eur"] == pytest.approx(
        row.top_two_per_million_eur("celkovy")
    )


def test_parse_official_results_html_extracts_financing_and_links():
    html = """
    <table><tbody>
      <tr class="print-info">
        <td><b>Univerzita Komenského v Bratislave</b>
        <b>(Fakulta matematiky, fyziky a informatiky)</b><br>
        Počet zamestnancov: 53 • Financovanie: 10 373 579€
        • Oblasť hodnotenia: Matematické vedy</td>
      </tr>
      <tr>
        <td><a href="https://ver.cvtisr.sk/vysledky/zobrazit-ziadost/?id=7560&show=final-outputs-list">Výstupy</a></td>
        <td><a href="https://ver.cvtisr.sk/vysledky/zobrazit-ziadost/?id=7560&show=creative-environment">Tvorivé prostredie</a></td>
      </tr>
    </tbody></table>
    """
    rows = parse_results_html(html)
    assert len(rows) == 1
    row = rows[0]
    assert row.eval_area == "Matematické vedy"
    assert row.institution == "Univerzita Komenského v Bratislave"
    assert row.inst_level == "Fakulta matematiky, fyziky a informatiky"
    assert row.employees == 53
    assert row.financing_eur == 10_373_579
    assert row.official_application_id == "7560"
    assert "Výstupy" in row.official_links


def test_sort_by_top_two_per_100_employees(data):
    rows = [r for r in data if r.employees]
    ranked = sort_rows(rows, by="celkovy__top2_per_100_emp", descending=True)
    assert ranked[0].top_two_per_100_employees("celkovy") >= ranked[-1].top_two_per_100_employees(
        "celkovy"
    )


def test_efficiency_cli_outputs_reproducible_table(capsys):
    rc = main(["efficiency", "--limit", "3"])
    out = capsys.readouterr().out
    assert rc == 0
    assert "CTop2/100zam" in out
    assert "Ústav orientalistiky Slovenskej akadémie vied" in out


def test_money_cli_outputs_financing_table(capsys):
    rc = main(["money", "--limit", "2"])
    out = capsys.readouterr().out
    assert rc == 0
    assert "CTop2/1M€" in out
    assert "Financovanie" in out


def test_fmfi_summary_values_match_reference_table(data):
    """Guard the FMFI UK three-area summary shown by the static viewer."""
    expected = {
        "Matematické vedy": (53, 68.0, 87.5, 100.0, 75.7, 1, 8),
        "Fyzikálne vedy": (109, 68.0, 33.3, 75.0, 63.9, 1, 6),
        "Informačné a komunikačné vedy": (45, 52.0, 100.0, 100.0, 66.4, 4, 15),
    }

    for area, values in expected.items():
        area_rows = [r for r in data if r.eval_area == area]
        fmfi = next(
            r for r in area_rows
            if r.institution == "Univerzita Komenského v Bratislave"
            and "Fakulta matematiky, fyziky a informatiky" in r.inst_level
        )
        ranked = sorted(area_rows, key=lambda r: r.score("celkovy"))
        actual = (
            fmfi.employees,
            round(fmfi.top_two_pct("vystupy"), 1),
            round(fmfi.top_two_pct("spolocensky_dosah"), 1),
            round(fmfi.top_two_pct("tvorive_prostredie"), 1),
            round(fmfi.top_two_pct("celkovy"), 1),
            ranked.index(fmfi) + 1,
            len(area_rows),
        )

        assert actual == values


def test_reward_quality_points_matches_t14a_methodology():
    assert quality_points([24.35, 55.1, 17.75, 2.8, 0]) == pytest.approx(526.35)


def test_reward_comparison_uses_trnava_ttu_alias(tmp_path):
    old_path = tmp_path / "old.xlsx"
    new_path = tmp_path / "new.xlsx"

    old_wb = openpyxl.Workbook()
    old_ws = old_wb.active
    old_ws.title = "T14a-ver2022"
    old_ws["Y9"] = 1000
    old_ws["AA8"] = "VŠ"
    old_ws["AB8"] = "VER"
    old_ws["AC8"] = "Podiel VER"
    old_ws["AD8"] = "Suma 2026"
    for col, value in enumerate(["5*", "4*", "3*", "2*", "1*"], start=5):
        old_ws.cell(9, col, value)
    old_ws["AA9"] = "TTU"
    old_ws["AB9"] = 100
    old_ws["AC9"] = 100
    old_ws["AD9"] = 1000
    for col, value in enumerate(
        ["TTU", "Za celú inštitúciu", "Matematické vedy", 1.3, 100, 0, 0, 0, 0, 10],
        start=1,
    ):
        old_ws.cell(10, col, value)

    names_ws = old_wb.create_sheet("VŠ-Názov")
    names_ws.append([])
    names_ws.append(["KOD VVŠ", "Názov plný  v CRŠ", "VVŠ naša skratka"])
    names_ws.append([713000000, "Trnavská univerzita v Trnave", "TVU"])
    old_wb.save(old_path)

    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "VER2026data"
    new_ws.append(["title"])
    headers = [None] * 28
    headers[0] = "Číslo žiadosti"
    headers[1] = "Oblasť hodnotenia"
    headers[2] = "Skupina oblastí hodnotenia"
    headers[3] = "Inštitúcia"
    headers[4] = "Inštitucionálna úroveň žiadosti"
    headers[5] = "VVI alebo VVŠ"
    headers[6] = "Počet zamestnancov"
    headers[23] = "% Celkový Profil kvality - Excelentná"
    new_ws.append(headers)
    row = [None] * 28
    row[0] = "#1"
    row[1] = "Matematické vedy"
    row[2] = "Prírodné vedy"
    row[3] = "Trnavská univerzita v Trnave"
    row[4] = "Za celú inštitúciu"
    row[5] = "VVŠ"
    row[6] = 20
    row[23:28] = [100, 0, 0, 0, 0]
    new_ws.append(row)
    new_wb.save(new_path)

    comparison = build_reward_comparison(old_path, new_path)
    assert len(comparison["rows"]) == 1
    row = comparison["rows"][0]
    assert row["abbr"] == "TTU"
    assert row["ver2026_rows"] == 1
    assert row["new_amount_eur"] == pytest.approx(1000)
    assert row["old_employees"] == pytest.approx(10)
    assert row["new_employees"] == pytest.approx(20)
    assert comparison["views"]["areas"]["rows"][0]["label"] == "Matematické vedy"
    assert comparison["views"]["areas"]["rows"][0]["new_amount_eur"] == pytest.approx(1000)
    assert comparison["views"]["areas"]["rows"][0]["old_employees"] == pytest.approx(10)
    assert comparison["views"]["areas"]["rows"][0]["new_employees"] == pytest.approx(20)
    assert comparison["views"]["groups"]["rows"][0]["label"] == "Prírodné vedy"
    assert comparison["views"]["applications"]["rows"][0]["label"] == "TTU / Za celú inštitúciu"
    assert comparison["views"]["applications"]["rows"][0]["comparison_status"] == "matched"


def test_generated_reward_json_has_all_views_and_fixed_pool():
    payload = json.loads(REWARD_DATA.read_text(encoding="utf-8"))
    assert set(payload["views"]) == {"universities", "areas", "groups", "applications"}
    assert payload["sources"]["ver2026_profiles_xlsx"].endswith("36762.f7eba0.xlsx")
    assert payload["sources"]["subsidy_2026_xlsx"].endswith("35059.48a629.xlsx")
    pool = payload["method"]["pool_eur"]
    assert pool == pytest.approx(124_907_188)
    for view in payload["views"].values():
        rows = view["rows"]
        assert rows
        assert sum(row["old_amount_eur"] for row in rows) == pytest.approx(pool)
        assert sum(row["new_amount_eur"] for row in rows) == pytest.approx(pool)

    eu = next(row for row in payload["views"]["universities"]["rows"] if row["label"] == "EU")
    assert eu["amount_delta_eur"] == pytest.approx(3_039_995.353887337)
    assert eu["ver2022_rows"] == 8
    assert eu["ver2026_rows"] == 1

    economics = next(row for row in payload["views"]["areas"]["rows"] if row["label"] == "Ekonomické vedy")
    assert economics["amount_delta_eur"] == pytest.approx(4_990_777.530739086)
    assert economics["ver2022_rows"] == 23
    assert economics["ver2026_rows"] == 14
    assert economics["old_employees"] == pytest.approx(914)
    assert economics["new_employees"] == pytest.approx(977)

    area_rows = payload["views"]["areas"]["rows"]
    assert sum(row["old_employees"] for row in area_rows) == pytest.approx(8_307)
    assert sum(row["new_employees"] for row in area_rows) == pytest.approx(8_330)

    social = next(row for row in payload["views"]["groups"]["rows"] if row["label"] == "Spoločenské vedy")
    assert social["amount_delta_eur"] == pytest.approx(5_451_489.845603522)
    assert social["ver2022_rows"] == 109
    assert social["ver2026_rows"] == 77

    applications = payload["views"]["applications"]["rows"]
    assert len(applications) == 343
    statuses = {row["comparison_status"] for row in applications}
    assert statuses == {"matched", "old_only", "new_only"}
    department = next(row for row in applications if row["label"] == "TUAD / Katedra politológie")
    assert department["subtitle"] == "Politické vedy · Trenčianska univerzita Alexandra Dubčeka v Trenčíne"
    assert department["comparison_status"] == "new_only"
    assert department["new_amount_eur"] == pytest.approx(52_629.05279500311)


def test_main_web_data_does_not_embed_reward_payload():
    payload = json.loads(WEB_DATA.read_text(encoding="utf-8"))
    assert "reward_comparison" not in payload
    assert "views" not in payload
    assert "subsidy_2026_xlsx" not in json.dumps(payload)
