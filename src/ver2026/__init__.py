"""VER 2026 data loading and sorting helpers."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl


# Profile column groups. The 5 levels map to scores 1..5 (1 = best, 5 = worst).
# Each profile yields a weighted score 1..5 and a "% at top level" metric.
PROFILE_GROUPS = {
    # slug : (label, column indices into the header row, level names)
    "vystupy": (
        "Profil kvality výstupov",
        list(range(8, 13)),
        ["Svetová", "Významná medzinárodná", "Medzinárodná", "Národná", "Nezaradené"],
    ),
    "spolocensky_dosah": (
        "Profil kvality spoločenského dosahu",
        list(range(13, 18)),
        ["Excelentná", "Veľmi dobrá", "Dobrá", "Uspokojivá", "Neuspokojivá"],
    ),
    "tvorive_prostredie": (
        "Profil kvality tvorivého prostredia",
        list(range(18, 23)),
        ["Excelentná", "Veľmi dobrá", "Dobrá", "Uspokojivá", "Neuspokojivá"],
    ),
    "celkovy": (
        "Celkový profil kvality",
        list(range(23, 28)),
        ["Excelentná", "Veľmi dobrá", "Dobrá", "Uspokojivá", "Neuspokojivá"],
    ),
}

# Base column names (the first 8 cols are the same on every row).
COL_REQUEST_ID = 0
COL_EVAL_AREA = 1
COL_EVAL_GROUP = 2
COL_INSTITUTION = 3
COL_INST_LEVEL = 4
COL_INST_TYPE = 5       # "VVI" (Slovak Academy institute) or "VVŠ" (university)
COL_EMPLOYEES = 6
COL_WOMEN = 7


@dataclass
class Institution:
    """A single evaluation request — a faculty, institute, or whole institution."""

    request_id: str
    eval_area: str
    eval_group: str
    institution: str
    inst_level: str
    inst_type: str          # "VVI" or "VVŠ"
    employees: int | None
    women: int | None
    profiles: dict[str, list[float]]   # slug -> [p1, p2, p3, p4, p5], each in 0..100

    def score(self, slug: str) -> float:
        """Weighted score in 1..5 (lower = better). NaN-safe: returns 5 if no data."""
        levels = self.profiles.get(slug) or []
        if not levels or sum(levels) == 0:
            return 5.0
        total = sum(levels)
        return sum((i + 1) * p for i, p in enumerate(levels)) / total

    def top_pct(self, slug: str) -> float:
        """Percentage of outputs / metrics at the top level (level 1)."""
        levels = self.profiles.get(slug) or []
        return levels[0] if levels else 0.0

    def top_two_pct(self, slug: str) -> float:
        """Percentage at top two levels (level 1 + level 2)."""
        levels = self.profiles.get(slug) or []
        return (levels[0] or 0) + (levels[1] or 0) if levels else 0.0

    def as_dict(self) -> dict:
        out = {
            "request_id": self.request_id,
            "eval_area": self.eval_area,
            "eval_group": self.eval_group,
            "institution": self.institution,
            "inst_level": self.inst_level,
            "inst_type": self.inst_type,
            "employees": self.employees,
            "women": self.women,
        }
        for slug, (label, _, levels) in PROFILE_GROUPS.items():
            pcts = self.profiles.get(slug) or [None] * 5
            for level, pct in zip(levels, pcts):
                out[f"{slug}__{level}"] = pct
            out[f"{slug}__score"] = self.score(slug)
            out[f"{slug}__top1"] = self.top_pct(slug)
            out[f"{slug}__top2"] = self.top_two_pct(slug)
        return out


def _to_int(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None


def _to_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load(path: str | Path) -> list[Institution]:
    """Load institutions from the official VER 2026 XLSX.

    The workbook has the title in row 1, headers in row 2, and one
    institution per row from row 3 onward.
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []

    out: list[Institution] = []
    for raw in rows[2:]:
        if not raw or not raw[COL_REQUEST_ID]:
            continue
        profiles: dict[str, list[float]] = {}
        for slug, (_, idxs, _) in PROFILE_GROUPS.items():
            profiles[slug] = [
                _to_float(raw[i]) if i < len(raw) else None for i in idxs
            ]
        out.append(
            Institution(
                request_id=str(raw[COL_REQUEST_ID]).strip(),
                eval_area=(raw[COL_EVAL_AREA] or "").strip(),
                eval_group=(raw[COL_EVAL_GROUP] or "").strip(),
                institution=(raw[COL_INSTITUTION] or "").strip(),
                inst_level=(raw[COL_INST_LEVEL] or "").strip(),
                inst_type=(raw[COL_INST_TYPE] or "").strip(),
                employees=_to_int(raw[COL_EMPLOYEES]),
                women=_to_int(raw[COL_WOMEN]),
                profiles=profiles,
            )
        )
    return out


def filter_rows(
    rows: Iterable[Institution],
    *,
    eval_area: str | None = None,
    eval_group: str | None = None,
    inst_type: str | None = None,
    min_employees: int | None = None,
    institution_contains: str | None = None,
) -> list[Institution]:
    """Apply common filters used by both the CLI and the web viewer."""
    out = []
    for r in rows:
        if eval_area and r.eval_area != eval_area:
            continue
        if eval_group and r.eval_group != eval_group:
            continue
        if inst_type and r.inst_type != inst_type:
            continue
        if min_employees is not None and (r.employees or 0) < min_employees:
            continue
        if institution_contains:
            if institution_contains.lower() not in r.institution.lower():
                continue
        out.append(r)
    return out


def sort_rows(
    rows: list[Institution],
    *,
    by: str = "vystupy__top1",
    descending: bool = True,
) -> list[Institution]:
    """Sort by a metric key. Lower score = better for `*__score` columns."""
    lower_is_better = by.endswith("__score")

    def keyfn(r: Institution) -> float:
        if "__" not in by:
            return 0.0
        slug, _, metric = by.partition("__")
        if metric == "score":
            return r.score(slug)
        if metric == "top1":
            return r.top_pct(slug)
        if metric == "top2":
            return r.top_two_pct(slug)
        # Specific level like "vystupy__Svetová" — find the index.
        levels = PROFILE_GROUPS[slug][2]
        if metric in levels:
            return (r.profiles.get(slug) or [0] * 5)[levels.index(metric)]
        return 0.0

    return sorted(rows, key=keyfn, reverse=descending and not lower_is_better)
