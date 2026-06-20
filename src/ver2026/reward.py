"""Reward allocation comparison for VER 2022 and VER 2026 profiles."""
from __future__ import annotations

from collections import defaultdict
from io import BytesIO
import json
from pathlib import Path
import re
from typing import Iterable
import urllib.request

import openpyxl


QUALITY_WEIGHTS = (8.0, 5.0, 3.0, 1.0, 0.0)
OLD_RESULTS_SHEET = "T14a-ver2022"
UNIVERSITY_NAMES_SHEET = "VŠ-Názov"
VER2026_RESULTS_PAGE_URL = "https://www.minedu.sk/vysledky-periodickeho-hodnotenia-ver-2026/"
VER2026_XLSX_URL = "https://www.minedu.sk/data/att/b00/36762.f7eba0.xlsx"
SUBSIDY_2026_PAGE_URL = (
    "https://www.minedu.sk/44358-sk/"
    "rozpis-dotacii-zo-statneho-rozpoctu-verejnym-vysokym-skolam-na-rok-2026/"
)
SUBSIDY_2026_XLSX_URL = "https://www.minedu.sk/data/att/7c5/35059.48a629.xlsx"
VER2022_METHOD_PDF_URL = (
    "https://www.minedu.sk/metodicke-usmernenie-k-pouzitiu-vysledkov-periodickeho-"
    "hodnotenia-ver-2022-pre-ucely-posudenia-kvality-urovne-tvorivej-cinnosti-pri-"
    "standardoch-pre-studijny-program/"
)

COL_NEW_AREA = 1
COL_NEW_GROUP = 2
COL_NEW_INSTITUTION = 3
COL_NEW_LEVEL = 4
COL_NEW_TYPE = 5
COL_NEW_EMPLOYEES = 6
COL_NEW_OVERALL_START = 23

AREA_ALIASES = {
    "stavebné inžinierstvo, architektúra a doprava": "stavebné inžinierstvo a architektúra",
}

UNIVERSITY_ALIASES = {
    "trnavská univerzita v trnave": "TTU",
}

OLD_ABBR_ALIASES = {
    "TUZV": "TUZVO",
    "UPJS": "UPJŠ",
    "VSMU": "VŠMU",
    "VSVU": "VŠVU",
    "ZU": "ŽU",
}


def normalize(value: object) -> str:
    """Normalize labels that come from different official workbooks."""
    value = str(value or "").replace("\xa0", " ")
    return re.sub(r"\s+", " ", value).strip().casefold()


def clean_label(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def canonical_area_key(area: object) -> str:
    key = normalize(area)
    return normalize(AREA_ALIASES.get(key, key))


def clean_component(value: object) -> str:
    label = clean_label(value)
    prefix = "Za súčasť inštitúcie - "
    if label.startswith(prefix):
        return label[len(prefix) :]
    return label


def application_key(abbr: str, area: object, component: object) -> str:
    return "|".join([abbr, canonical_area_key(area), normalize(clean_component(component))])


def to_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def quality_points(levels: Iterable[object]) -> float:
    """Return 8A + 5B + 3C + D for a five-level quality profile."""
    values = [to_float(value) or 0.0 for value in levels]
    values += [0.0] * (len(QUALITY_WEIGHTS) - len(values))
    return sum(value * weight for value, weight in zip(values, QUALITY_WEIGHTS))


def weighted_volume(levels: Iterable[object], employees: object, cost_coefficient: object) -> float:
    """Quality points times employees times area cost coefficient.

    Source profiles are percentages, so the volume is divided by 100. The
    division has no effect on shares, but keeps the scale comparable to T14a.
    """
    return (
        quality_points(levels)
        * (to_float(employees) or 0.0)
        * (to_float(cost_coefficient) or 0.0)
        / 100.0
    )


def _load_workbook(path_or_url: str | Path):
    value = str(path_or_url)
    if value.startswith(("http://", "https://")):
        with urllib.request.urlopen(value, timeout=60) as response:
            return openpyxl.load_workbook(BytesIO(response.read()), read_only=True, data_only=True)
    return openpyxl.load_workbook(path_or_url, read_only=True, data_only=True)


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise ValueError(message)


def _validate_old_sheet(ws) -> None:
    _require(ws.title == OLD_RESULTS_SHEET, f"Expected old sheet {OLD_RESULTS_SHEET!r}, got {ws.title!r}")
    headers = [ws.cell(8, col).value for col in range(27, 31)]
    _require(headers == ["VŠ", "VER", "Podiel VER", "Suma 2026"], f"Unexpected T14a summary headers: {headers!r}")
    profile_headers = [ws.cell(9, col).value for col in range(5, 10)]
    _require(profile_headers == ["5*", "4*", "3*", "2*", "1*"], f"Unexpected T14a profile headers: {profile_headers!r}")
    _require(to_float(ws["Y9"].value) is not None, "Missing T14a reward pool in cell Y9")


def _validate_new_sheet(ws) -> None:
    headers = [ws.cell(2, col).value for col in range(1, 29)]
    expected = {
        0: "Číslo žiadosti",
        1: "Oblasť hodnotenia",
        2: "Skupina oblastí hodnotenia",
        3: "Inštitúcia",
        4: "Inštitucionálna úroveň žiadosti",
        5: "VVI alebo VVŠ",
        6: "Počet zamestnancov",
    }
    for index, value in expected.items():
        _require(headers[index] == value, f"Unexpected VER 2026 header at column {index + 1}: {headers[index]!r}")
    _require(
        str(headers[COL_NEW_OVERALL_START] or "").startswith("% Celkový Profil kvality"),
        f"Unexpected VER 2026 overall-profile header: {headers[COL_NEW_OVERALL_START]!r}",
    )


def _load_old_summary(ws) -> tuple[float, dict[str, dict[str, float]]]:
    summary: dict[str, dict[str, float]] = {}
    for row in ws.iter_rows(values_only=True):
        abbr = row[26] if len(row) > 26 else None
        old_volume = to_float(row[27] if len(row) > 27 else None)
        old_share = to_float(row[28] if len(row) > 28 else None)
        old_amount = to_float(row[29] if len(row) > 29 else None)
        if not isinstance(abbr, str) or old_volume is None or old_share is None or old_amount is None:
            continue
        summary[abbr.strip()] = {
            "old_weighted_volume": old_volume,
            "old_share_pct": old_share,
            "old_amount_eur": old_amount,
        }

    pool = to_float(ws["Y9"].value) or sum(row["old_amount_eur"] for row in summary.values())
    return pool, summary


def _load_area_costs(ws) -> dict[str, float]:
    costs: dict[str, float] = {}
    for row in ws.iter_rows(min_row=10, values_only=True):
        institution = row[0] if len(row) > 0 else None
        area = row[2] if len(row) > 2 else None
        coefficient = to_float(row[3] if len(row) > 3 else None)
        if not institution or not area or normalize(area) == "priemer" or coefficient is None:
            continue
        costs[canonical_area_key(area)] = coefficient
    return costs


def _area_cost(area: object, costs: dict[str, float]) -> float:
    key = canonical_area_key(area)
    if key in costs:
        return costs[key]
    raise KeyError(f"Missing cost coefficient for evaluation area {area!r}")


def _load_university_names(wb) -> tuple[dict[str, str], dict[str, str]]:
    ws = wb[UNIVERSITY_NAMES_SHEET]
    name_to_abbr: dict[str, str] = {}
    abbr_to_name: dict[str, str] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        full_name = row[1] if len(row) > 1 else None
        abbr = row[2] if len(row) > 2 else None
        if not full_name or not abbr:
            continue
        normalized = normalize(full_name)
        canonical = UNIVERSITY_ALIASES.get(normalized, str(abbr).strip())
        name_to_abbr[normalized] = canonical
        abbr_to_name[canonical] = str(full_name).replace("\xa0", " ").strip()
    return name_to_abbr, abbr_to_name


def _old_university_row_counts(ws) -> dict[str, int]:
    counts: dict[str, int] = defaultdict(int)
    for row in ws.iter_rows(min_row=10, values_only=True):
        abbr = row[0] if len(row) > 0 else None
        area = row[2] if len(row) > 2 else None
        if not abbr or not area or normalize(area) == "priemer":
            continue
        counts[OLD_ABBR_ALIASES.get(str(abbr).strip(), str(abbr).strip())] += 1
    return dict(counts)


def _old_university_employees(ws) -> dict[str, float]:
    employees: dict[str, float] = defaultdict(float)
    for row in ws.iter_rows(min_row=10, values_only=True):
        abbr = row[0] if len(row) > 0 else None
        area = row[2] if len(row) > 2 else None
        if not abbr or not area or normalize(area) == "priemer":
            continue
        employees[OLD_ABBR_ALIASES.get(str(abbr).strip(), str(abbr).strip())] += (
            to_float(row[9] if len(row) > 9 else None) or 0.0
        )
    return dict(employees)


def _new_dimensions(
    path: str | Path,
    *,
    area_costs: dict[str, float],
    name_to_abbr: dict[str, str],
) -> tuple[
    dict[str, dict[str, float]],
    dict[str, dict[str, int]],
    dict[str, dict[str, float]],
    dict[str, dict[str, str]],
    dict[str, dict[str, str]],
    dict[str, str],
]:
    wb = _load_workbook(path)
    ws = wb["VER2026data"] if "VER2026data" in wb.sheetnames else wb[wb.sheetnames[0]]
    _validate_new_sheet(ws)
    volumes = {
        "universities": defaultdict(float),
        "areas": defaultdict(float),
        "groups": defaultdict(float),
        "applications": defaultdict(float),
    }
    row_counts = {
        "universities": defaultdict(int),
        "areas": defaultdict(int),
        "groups": defaultdict(int),
        "applications": defaultdict(int),
    }
    employees = {
        "universities": defaultdict(float),
        "areas": defaultdict(float),
        "groups": defaultdict(float),
        "applications": defaultdict(float),
    }
    labels = {
        "universities": {},
        "areas": {},
        "groups": {},
        "applications": {},
    }
    subtitles = {
        "universities": {},
        "areas": {},
        "groups": {},
        "applications": {},
    }
    area_to_group: dict[str, str] = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0] or row[COL_NEW_TYPE] != "VVŠ":
            continue
        institution = row[COL_NEW_INSTITUTION]
        abbr = name_to_abbr.get(normalize(institution))
        if not abbr:
            raise KeyError(f"Missing university abbreviation for {institution!r}")
        area_key = canonical_area_key(row[COL_NEW_AREA])
        group_key = normalize(row[COL_NEW_GROUP])
        area_label = clean_label(row[COL_NEW_AREA])
        group_label = clean_label(row[COL_NEW_GROUP])
        component = clean_component(row[COL_NEW_LEVEL])
        employee_count = to_float(row[COL_NEW_EMPLOYEES]) or 0.0
        levels = row[COL_NEW_OVERALL_START : COL_NEW_OVERALL_START + 5]
        volume = weighted_volume(levels, employee_count, _area_cost(row[COL_NEW_AREA], area_costs))

        volumes["universities"][abbr] += volume
        row_counts["universities"][abbr] += 1
        employees["universities"][abbr] += employee_count
        labels["universities"][abbr] = abbr
        subtitles["universities"][abbr] = clean_label(institution)

        volumes["areas"][area_key] += volume
        row_counts["areas"][area_key] += 1
        employees["areas"][area_key] += employee_count
        labels["areas"][area_key] = area_label
        subtitles["areas"][area_key] = group_label
        area_to_group[area_key] = group_label

        volumes["groups"][group_key] += volume
        row_counts["groups"][group_key] += 1
        employees["groups"][group_key] += employee_count
        labels["groups"][group_key] = group_label
        subtitles["groups"][group_key] = "Skupina oblastí hodnotenia"

        app_key = application_key(abbr, row[COL_NEW_AREA], component)
        volumes["applications"][app_key] += volume
        row_counts["applications"][app_key] += 1
        employees["applications"][app_key] += employee_count
        labels["applications"][app_key] = f"{abbr} / {component}"
        subtitles["applications"][app_key] = f"{area_label} · {clean_label(institution)}"

    return (
        {name: dict(items) for name, items in volumes.items()},
        {name: dict(items) for name, items in row_counts.items()},
        {name: dict(items) for name, items in employees.items()},
        labels,
        subtitles,
        area_to_group,
    )


def _old_application_volumes(
    ws,
    *,
    abbr_to_name: dict[str, str],
) -> tuple[dict[str, float], dict[str, int], dict[str, float], dict[str, str], dict[str, str]]:
    volumes: dict[str, float] = defaultdict(float)
    row_counts: dict[str, int] = defaultdict(int)
    employees: dict[str, float] = defaultdict(float)
    labels: dict[str, str] = {}
    subtitles: dict[str, str] = {}

    for row in ws.iter_rows(min_row=10, values_only=True):
        abbr = row[0] if len(row) > 0 else None
        component = row[1] if len(row) > 1 else None
        area = row[2] if len(row) > 2 else None
        if not abbr or not component or not area or normalize(area) == "priemer":
            continue
        canonical_abbr = OLD_ABBR_ALIASES.get(str(abbr).strip(), str(abbr).strip())
        component_label = clean_component(component)
        key = application_key(canonical_abbr, area, component_label)
        employee_count = to_float(row[9] if len(row) > 9 else None) or 0.0
        volume = weighted_volume(row[4:9], employee_count, row[3] if len(row) > 3 else None)

        volumes[key] += volume
        row_counts[key] += 1
        employees[key] += employee_count
        labels[key] = f"{canonical_abbr} / {component_label}"
        subtitles[key] = f"{clean_label(area)} · {abbr_to_name.get(canonical_abbr, canonical_abbr)}"

    return dict(volumes), dict(row_counts), dict(employees), labels, subtitles


def _old_category_volumes(
    ws,
    area_to_group: dict[str, str],
) -> tuple[dict[str, dict[str, float]], dict[str, dict[str, int]], dict[str, dict[str, float]], dict[str, dict[str, str]]]:
    volumes = {
        "areas": defaultdict(float),
        "groups": defaultdict(float),
    }
    row_counts = {
        "areas": defaultdict(int),
        "groups": defaultdict(int),
    }
    employees = {
        "areas": defaultdict(float),
        "groups": defaultdict(float),
    }
    labels = {
        "areas": {},
        "groups": {},
    }
    missing_groups = []

    for row in ws.iter_rows(min_row=10, values_only=True):
        institution = row[0] if len(row) > 0 else None
        area = row[2] if len(row) > 2 else None
        if not institution or not area or normalize(area) == "priemer":
            continue
        area_key = canonical_area_key(area)
        group_label = area_to_group.get(area_key, "")
        group_key = normalize(group_label)
        if not group_key:
            missing_groups.append(clean_label(area))
            continue
        employee_count = to_float(row[9] if len(row) > 9 else None) or 0.0
        volume = weighted_volume(row[4:9], employee_count, row[3] if len(row) > 3 else None)

        volumes["areas"][area_key] += volume
        row_counts["areas"][area_key] += 1
        employees["areas"][area_key] += employee_count
        labels["areas"][area_key] = clean_label(area)

        volumes["groups"][group_key] += volume
        row_counts["groups"][group_key] += 1
        employees["groups"][group_key] += employee_count
        labels["groups"][group_key] = group_label

    _require(
        not missing_groups,
        "Missing VER 2026 group mapping for old T14a areas: " + ", ".join(sorted(set(missing_groups))),
    )

    return (
        {name: dict(items) for name, items in volumes.items()},
        {name: dict(items) for name, items in row_counts.items()},
        {name: dict(items) for name, items in employees.items()},
        labels,
    )


def _metrics_from_volumes(volumes: dict[str, float], pool_eur: float) -> dict[str, dict[str, float]]:
    total = sum(volumes.values())
    if total <= 0:
        return {}
    return {
        key: {
            "old_weighted_volume": volume,
            "old_share_pct": volume / total * 100.0,
            "old_amount_eur": pool_eur * volume / total,
        }
        for key, volume in volumes.items()
    }


def _comparison_rows(
    *,
    old_metrics: dict[str, dict[str, float]],
    new_volumes: dict[str, float],
    pool_eur: float,
    labels: dict[str, str],
    subtitles: dict[str, str] | None = None,
    old_row_counts: dict[str, int] | None = None,
    new_row_counts: dict[str, int] | None = None,
    old_employees: dict[str, float] | None = None,
    new_employees: dict[str, float] | None = None,
    abbr_to_name: dict[str, str] | None = None,
    require_matching_keys: bool = True,
) -> list[dict]:
    new_total = sum(new_volumes.values())
    if new_total <= 0:
        raise ValueError("VER 2026 weighted volume is zero; cannot compute shares")
    old_keys = set(old_metrics)
    new_keys = set(new_volumes)
    if require_matching_keys:
        _require(
            old_keys == new_keys,
            "Old/new reward keys differ; old-only="
            + repr(sorted(old_keys - new_keys))
            + ", new-only="
            + repr(sorted(new_keys - old_keys)),
        )

    rows = []
    for key in sorted(set(old_metrics) | set(new_volumes), key=lambda item: labels.get(item, item)):
        old = old_metrics.get(
            key,
            {
                "old_weighted_volume": 0.0,
                "old_share_pct": 0.0,
                "old_amount_eur": 0.0,
            },
        )
        new_weighted_volume = new_volumes.get(key, 0.0)
        new_share_pct = new_weighted_volume / new_total * 100.0
        new_amount_eur = pool_eur * new_share_pct / 100.0
        old_amount_eur = old["old_amount_eur"]
        old_share_pct = old["old_share_pct"]
        amount_delta_eur = new_amount_eur - old_amount_eur
        row = {
            "key": key,
            "label": labels.get(key, key),
            "subtitle": (subtitles or {}).get(key, ""),
            "old_weighted_volume": old["old_weighted_volume"],
            "new_weighted_volume": new_weighted_volume,
            "old_share_pct": old_share_pct,
            "new_share_pct": new_share_pct,
            "share_delta_pct": new_share_pct - old_share_pct,
            "old_amount_eur": old_amount_eur,
            "new_amount_eur": new_amount_eur,
            "amount_delta_eur": amount_delta_eur,
            "amount_delta_pct": (amount_delta_eur / old_amount_eur * 100.0) if old_amount_eur else None,
            "ver2022_rows": (old_row_counts or {}).get(key, 0),
            "ver2026_rows": (new_row_counts or {}).get(key, 0),
            "old_employees": (old_employees or {}).get(key, 0.0),
            "new_employees": (new_employees or {}).get(key, 0.0),
            "comparison_status": "matched"
            if key in old_metrics and key in new_volumes
            else "new_only"
            if key in new_volumes
            else "old_only",
        }
        if abbr_to_name is not None:
            row["abbr"] = key
            row["institution"] = abbr_to_name.get(key, key)
            row["label"] = key
            row["subtitle"] = abbr_to_name.get(key, key)
        rows.append(row)

    rows.sort(key=lambda row: row["amount_delta_eur"], reverse=True)
    return rows


def build_reward_comparison(ver2022_path: str | Path, ver2026_path: str | Path) -> dict:
    """Build a VER 2022 -> VER 2026 reward allocation comparison.

    The university view uses the old official T14a summary. Category views use
    row-level T14a data. Every view keeps the same total pool and recomputes
    VER 2026 shares from overall quality profiles with the same 8/5/3/1/0
    weights and area cost coefficients.
    """
    old_wb = _load_workbook(ver2022_path)
    old_ws = old_wb[OLD_RESULTS_SHEET]
    _validate_old_sheet(old_ws)
    pool_eur, old_summary = _load_old_summary(old_ws)
    old_university_counts = _old_university_row_counts(old_ws)
    old_university_employees = _old_university_employees(old_ws)
    area_costs = _load_area_costs(old_ws)
    name_to_abbr, abbr_to_name = _load_university_names(old_wb)
    new_volumes, new_row_counts, new_employees, new_labels, new_subtitles, area_to_group = _new_dimensions(
        ver2026_path,
        area_costs=area_costs,
        name_to_abbr=name_to_abbr,
    )
    old_category_volumes, old_category_counts, old_category_employees, old_category_labels = _old_category_volumes(
        old_ws,
        area_to_group,
    )
    old_application_volumes, old_application_counts, old_application_employees, old_application_labels, old_application_subtitles = (
        _old_application_volumes(old_ws, abbr_to_name=abbr_to_name)
    )

    university_rows = _comparison_rows(
        old_metrics=old_summary,
        new_volumes=new_volumes["universities"],
        pool_eur=pool_eur,
        labels=new_labels["universities"],
        subtitles=new_subtitles["universities"],
        old_row_counts=old_university_counts,
        new_row_counts=new_row_counts["universities"],
        old_employees=old_university_employees,
        new_employees=new_employees["universities"],
        abbr_to_name=abbr_to_name,
    )
    area_labels = {**old_category_labels["areas"], **new_labels["areas"]}
    area_rows = _comparison_rows(
        old_metrics=_metrics_from_volumes(old_category_volumes["areas"], pool_eur),
        new_volumes=new_volumes["areas"],
        pool_eur=pool_eur,
        labels=area_labels,
        subtitles=new_subtitles["areas"],
        old_row_counts=old_category_counts["areas"],
        new_row_counts=new_row_counts["areas"],
        old_employees=old_category_employees["areas"],
        new_employees=new_employees["areas"],
    )
    group_labels = {**old_category_labels["groups"], **new_labels["groups"]}
    group_rows = _comparison_rows(
        old_metrics=_metrics_from_volumes(old_category_volumes["groups"], pool_eur),
        new_volumes=new_volumes["groups"],
        pool_eur=pool_eur,
        labels=group_labels,
        subtitles=new_subtitles["groups"],
        old_row_counts=old_category_counts["groups"],
        new_row_counts=new_row_counts["groups"],
        old_employees=old_category_employees["groups"],
        new_employees=new_employees["groups"],
    )
    application_labels = {**old_application_labels, **new_labels["applications"]}
    application_subtitles = {**old_application_subtitles, **new_subtitles["applications"]}
    application_rows = _comparison_rows(
        old_metrics=_metrics_from_volumes(old_application_volumes, pool_eur),
        new_volumes=new_volumes["applications"],
        pool_eur=pool_eur,
        labels=application_labels,
        subtitles=application_subtitles,
        old_row_counts=old_application_counts,
        new_row_counts=new_row_counts["applications"],
        old_employees=old_application_employees,
        new_employees=new_employees["applications"],
        require_matching_keys=False,
    )

    return {
        "sources": {
            "ver2026_results_page": VER2026_RESULTS_PAGE_URL,
            "ver2026_profiles_xlsx": VER2026_XLSX_URL,
            "subsidy_2026_page": SUBSIDY_2026_PAGE_URL,
            "subsidy_2026_xlsx": SUBSIDY_2026_XLSX_URL,
            "ver2022_methodology_pdf": VER2022_METHOD_PDF_URL,
            "ver2022_sheet": OLD_RESULTS_SHEET,
        },
        "method": {
            "quality_weights": list(QUALITY_WEIGHTS),
            "pool_eur": pool_eur,
            "old_total_weighted_volume": sum(row["old_weighted_volume"] for row in old_summary.values()),
            "new_total_weighted_volume": sum(new_volumes["universities"].values()),
            "old_source_sheet": OLD_RESULTS_SHEET,
            "basis": "Same total pool as T14a; VER 2026 shares recomputed from overall quality profiles for VVŠ only.",
            "quality_formula": (
                "A/B/C/D/E are the five percentages in the overall quality profile: "
                "Excellent, Very good, Good, Satisfactory, Unsatisfactory. "
                "The weighted quality is 8*A + 5*B + 3*C + 1*D + 0*E."
            ),
            "application_view_note": (
                "Application rows are matched by university abbreviation, evaluation area, and component label. "
                "Rows that exist only in one source are kept as old_only or new_only."
            ),
        },
        "cost_coefficients": {
            area: coefficient
            for area, coefficient in sorted(area_costs.items())
        },
        "views": {
            "universities": {
                "label": "VVŠ",
                "unit_label": "Verejné vysoké školy",
                "rows": university_rows,
            },
            "areas": {
                "label": "Oblasti hodnotenia",
                "unit_label": "Oblasti hodnotenia",
                "rows": area_rows,
            },
            "groups": {
                "label": "Skupiny oblastí",
                "unit_label": "Skupiny oblastí hodnotenia",
                "rows": group_rows,
            },
            "applications": {
                "label": "Žiadosti / súčasti",
                "unit_label": "Riadky žiadostí a súčastí",
                "rows": application_rows,
            },
        },
        "rows": university_rows,
    }


def write_reward_comparison(ver2022_path: str | Path, ver2026_path: str | Path, out_path: str | Path) -> dict:
    comparison = build_reward_comparison(ver2022_path, ver2026_path)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(comparison, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return comparison
